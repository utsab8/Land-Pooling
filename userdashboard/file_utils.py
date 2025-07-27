import os
import json
import csv
import zipfile
import tempfile
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon, LineString, mapping
from shapely.wkt import loads
import pyogrio
import xml.etree.ElementTree as ET
from io import StringIO, BytesIO
from django.conf import settings
from django.http import HttpResponse
from django.core.files.storage import default_storage
import logging
from decimal import Decimal
import re

logger = logging.getLogger(__name__)

class FileValidator:
    """Comprehensive file validation utility"""
    
    ALLOWED_EXTENSIONS = {
        'kml': ['.kml'],
        'csv': ['.csv'],
        'shapefile': ['.shp', '.zip'],
        'geojson': ['.geojson', '.json'],
        'excel': ['.xlsx', '.xls'],
        'pdf': ['.pdf'],
        'image': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp'],
        'other': ['.txt', '.doc', '.docx', '.rtf']
    }
    
    MAX_FILE_SIZES = {
        'kml': 50,  # 50MB
        'csv': 100,  # 100MB
        'shapefile': 200,  # 200MB
        'geojson': 100,  # 100MB
        'excel': 50,  # 50MB
        'pdf': 100,  # 100MB
        'image': 20,  # 20MB
        'other': 50  # 50MB
    }
    
    @classmethod
    def detect_file_type(cls, file_obj, filename):
        """Detect file type based on extension and content"""
        extension = os.path.splitext(filename.lower())[1]
        
        # First check by extension
        for file_type, extensions in cls.ALLOWED_EXTENSIONS.items():
            if extension in extensions:
                return file_type
        
        # If extension not found, try content-based detection
        try:
            # Read first few bytes for basic detection
            file_obj.seek(0)
            header = file_obj.read(1024)
            file_obj.seek(0)
            
            # Basic content detection without magic
            if header.startswith(b'%PDF'):
                return 'pdf'
            elif header.startswith(b'\xff\xd8\xff'):  # JPEG
                return 'image'
            elif header.startswith(b'\x89PNG\r\n\x1a\n'):  # PNG
                return 'image'
            elif header.startswith(b'GIF87a') or header.startswith(b'GIF89a'):  # GIF
                return 'image'
            elif header.startswith(b'PK'):  # ZIP
                return 'shapefile'
            elif b'<?xml' in header and b'kml' in header.lower():
                return 'kml'
            elif b'{' in header and b'"type"' in header.lower():
                return 'geojson'
            elif b',' in header and b'\n' in header:
                return 'csv'
                
        except Exception as e:
            logger.warning(f"Error detecting file type for {filename}: {e}")
        
        return 'other'
    
    @classmethod
    def validate_file(cls, file_obj, filename, file_type=None):
        """Comprehensive file validation"""
        errors = []
        warnings = []
        
        # Get file size in MB
        file_obj.seek(0, 2)  # Seek to end
        file_size_mb = file_obj.tell() / (1024 * 1024)
        file_obj.seek(0)  # Reset to beginning
        
        # Detect file type if not provided
        if not file_type:
            file_type = cls.detect_file_type(file_obj, filename)
        
        # Validate file size
        max_size = cls.MAX_FILE_SIZES.get(file_type, 50)
        if file_size_mb > max_size:
            errors.append(f"File size ({file_size_mb:.2f}MB) exceeds maximum allowed size ({max_size}MB)")
        
        # Validate extension
        extension = os.path.splitext(filename.lower())[1]
        allowed_extensions = cls.ALLOWED_EXTENSIONS.get(file_type, [])
        if allowed_extensions and extension not in allowed_extensions:
            errors.append(f"File extension '{extension}' is not allowed for {file_type} files")
        
        # Content-specific validation
        content_errors = cls._validate_content(file_obj, file_type, filename)
        errors.extend(content_errors)
        
        return {
            'is_valid': len(errors) == 0,
            'file_type': file_type,
            'file_size_mb': file_size_mb,
            'errors': errors,
            'warnings': warnings
        }
    
    @classmethod
    def _validate_content(cls, file_obj, file_type, filename):
        """Validate file content based on type"""
        errors = []
        
        try:
            if file_type == 'kml':
                errors.extend(cls._validate_kml(file_obj))
            elif file_type == 'csv':
                errors.extend(cls._validate_csv(file_obj))
            elif file_type == 'excel':
                errors.extend(cls._validate_excel(file_obj))
            elif file_type == 'image':
                errors.extend(cls._validate_image(file_obj))
            elif file_type == 'pdf':
                errors.extend(cls._validate_pdf(file_obj))
            elif file_type == 'shapefile':
                errors.extend(cls._validate_shapefile(file_obj, filename))
            elif file_type == 'geojson':
                errors.extend(cls._validate_geojson(file_obj))
                
        except Exception as e:
            errors.append(f"Error validating {file_type} content: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_kml(cls, file_obj):
        """Validate KML file content"""
        errors = []
        try:
            tree = ET.parse(file_obj)
            root = tree.getroot()
            
            # Check for KML namespace
            if not root.tag.endswith('kml'):
                errors.append("Invalid KML file: missing or incorrect root element")
            
            # Check for placemarks
            placemarks = root.findall('.//{http://www.opengis.net/kml/2.2}Placemark')
            if not placemarks:
                errors.append("KML file contains no placemarks")
                
        except ET.ParseError:
            errors.append("Invalid KML file: XML parsing error")
        except Exception as e:
            errors.append(f"KML validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_csv(cls, file_obj):
        """Validate CSV file content"""
        errors = []
        try:
            # Try to read first few lines
            content = file_obj.read(1024).decode('utf-8')
            file_obj.seek(0)
            
            lines = content.split('\n')
            if len(lines) < 2:
                errors.append("CSV file must contain at least header and one data row")
            
            # Check if it's actually comma-separated
            first_line = lines[0]
            if ',' not in first_line:
                errors.append("File does not appear to be comma-separated")
                
        except UnicodeDecodeError:
            errors.append("CSV file must be UTF-8 encoded")
        except Exception as e:
            errors.append(f"CSV validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_excel(cls, file_obj):
        """Validate Excel file content"""
        errors = []
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_obj, read_only=True)
            if not workbook.sheetnames:
                errors.append("Excel file contains no worksheets")
            file_obj.seek(0)
        except Exception as e:
            errors.append(f"Excel validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_image(cls, file_obj):
        """Validate image file content"""
        errors = []
        try:
            from PIL import Image
            image = Image.open(file_obj)
            image.verify()
            file_obj.seek(0)
        except Exception as e:
            errors.append(f"Image validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_pdf(cls, file_obj):
        """Validate PDF file content"""
        errors = []
        try:
            # Read first few bytes to check PDF signature
            header = file_obj.read(4)
            file_obj.seek(0)
            if header != b'%PDF':
                errors.append("Invalid PDF file: missing PDF signature")
        except Exception as e:
            errors.append(f"PDF validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_shapefile(cls, file_obj, filename):
        """Validate shapefile content"""
        errors = []
        try:
            if filename.lower().endswith('.zip'):
                # Check if it's a valid ZIP containing shapefile components
                with zipfile.ZipFile(file_obj, 'r') as zip_file:
                    shp_files = [f for f in zip_file.namelist() if f.lower().endswith('.shp')]
                    if not shp_files:
                        errors.append("ZIP file does not contain .shp file")
            else:
                # Direct .shp file
                if not filename.lower().endswith('.shp'):
                    errors.append("Shapefile must be .shp or .zip format")
            file_obj.seek(0)
        except Exception as e:
            errors.append(f"Shapefile validation error: {str(e)}")
        
        return errors
    
    @classmethod
    def _validate_geojson(cls, file_obj):
        """Validate GeoJSON file content"""
        errors = []
        try:
            content = file_obj.read().decode('utf-8')
            file_obj.seek(0)
            
            data = json.loads(content)
            if 'type' not in data:
                errors.append("Invalid GeoJSON: missing 'type' field")
            elif data['type'] not in ['Feature', 'FeatureCollection', 'Geometry']:
                errors.append("Invalid GeoJSON: unsupported type")
                
        except json.JSONDecodeError:
            errors.append("Invalid GeoJSON: JSON parsing error")
        except UnicodeDecodeError:
            errors.append("GeoJSON file must be UTF-8 encoded")
        except Exception as e:
            errors.append(f"GeoJSON validation error: {str(e)}")
        
        return errors

class FileProcessor:
    """Comprehensive file processor for KML, CSV, and Shapefile handling"""
    
    def __init__(self, file_upload):
        self.file_upload = file_upload
        self.file_path = file_upload.file.path
        self.file_type = file_upload.file_type
        
    def process_file(self):
        """Process uploaded file based on its type"""
        try:
            if self.file_type == 'kml':
                return self._process_kml()
            elif self.file_type == 'csv':
                return self._process_csv()
            elif self.file_type == 'shapefile':
                return self._process_shapefile()
            else:
                raise ValueError(f"Unsupported file type: {self.file_type}")
        except Exception as e:
            logger.error(f"Error processing file {self.file_upload.id}: {e}")
            raise
    
    def _process_kml(self):
        """Process KML file and extract data"""
        from .kml_utils import KMLParser
        
        parser = KMLParser(self.file_path)
        parsed_data = parser.parse_kml()
        
        if not parsed_data:
            raise ValueError("No valid placemarks found in KML file")
        
        # Save to KMLData model
        from .models import KMLData
        kml_data_objects = []
        
        for data in parsed_data:
            kml_data = KMLData.objects.create(
                kml_file=self.file_upload.kml_files.first() if hasattr(self.file_upload, 'kml_files') else None,
                placemark_name=data.get('placemark_name', ''),
                kitta_number=data.get('kitta_number', ''),
                owner_name=data.get('owner_name', ''),
                geometry_type=data.get('geometry_type', ''),
                coordinates=data.get('coordinates', ''),
                area_hectares=data.get('area_hectares'),
                area_sqm=data.get('area_sqm'),
                description=data.get('description', ''),
                extended_data=data.get('extended_data', {}),
                time_begin=data.get('time_begin', ''),
                time_end=data.get('time_end', ''),
                time_when=data.get('time_when', ''),
                altitude=data.get('altitude', ''),
                altitude_mode=data.get('altitude_mode', ''),
                tessellate=data.get('tessellate', ''),
                extrude=data.get('extrude', ''),
                style_id=data.get('style_id', ''),
                style_url=data.get('style_url', ''),
                address=data.get('address', ''),
                country_code=data.get('country_code', ''),
                administrative_area=data.get('administrative_area', ''),
                sub_administrative_area=data.get('sub_administrative_area', ''),
                locality=data.get('locality', ''),
                sub_locality=data.get('sub_locality', ''),
                thoroughfare=data.get('thoroughfare', ''),
                postal_code=data.get('postal_code', ''),
                phone_number=data.get('phone_number', ''),
                snippet=data.get('snippet', ''),
                visibility=data.get('visibility', ''),
                open_status=data.get('open', ''),
                atom_author=data.get('atom_author', ''),
                atom_link=data.get('atom_link', ''),
                xal_address_details=data.get('xal_address_details', ''),
            )
            kml_data_objects.append(kml_data)
        
        # Update file metadata
        self.file_upload.geometry_type = self._get_geometry_type_from_data(parsed_data)
        self.file_upload.feature_count = len(parsed_data)
        self.file_upload.bounds = self._calculate_bounds(parsed_data)
        self.file_upload.status = 'completed'
        self.file_upload.save()
        
        return {
            'success': True,
            'data_count': len(parsed_data),
            'geometry_type': self.file_upload.geometry_type,
            'bounds': self.file_upload.bounds
        }
    
    def _process_csv(self):
        """Process CSV file and extract data, supporting polygons via 'Coordinates' column"""
        try:
            import ast
            # Read CSV file
            df = pd.read_csv(self.file_path)
            print('CSV Columns:', df.columns)
            if df.empty:
                print('CSV is empty')
                raise ValueError("CSV file is empty")

            # Detect coordinate columns
            lat_col, lon_col = self._detect_coordinate_columns(df)
            print('Detected lat_col:', lat_col, 'lon_col:', lon_col)
            if not lat_col and not lon_col and 'coordinates' not in [c.lower() for c in df.columns]:
                print('No coordinates column found')
                raise ValueError("Could not detect latitude and longitude columns or 'Coordinates' column")

            from .models import CSVData
            csv_data_objects = []

            for index, row in df.iterrows():
                try:
                    print(f'Processing row {index+1}')
                    if lat_col and lon_col:
                        lat = float(row[lat_col])
                        lon = float(row[lon_col])
                        geometry_type = 'Point'
                        coords = [lon, lat]
                    elif 'coordinates' in [c.lower() for c in df.columns]:
                        # Parse the first coordinate pair from the string
                        coords_val = row[[c for c in df.columns if c.lower() == 'coordinates'][0]]
                        coords_list = ast.literal_eval(coords_val)
                        if not coords_list or not isinstance(coords_list, list):
                            print('Invalid coordinates list:', coords_list)
                            continue
                        lon, lat = coords_list[0]
                        geometry_type = 'Polygon'
                        coords = coords_list
                    else:
                        print('No valid coordinates for row', index+1)
                        continue

                    # Sanitize row data for JSONField
                    row_data = {}
                    for k, v in row.items():
                        if pd.isna(v):
                            row_data[k] = None
                        elif isinstance(v, float) and (v == float('inf') or v == float('-inf')):
                            row_data[k] = None
                        elif k.lower() == 'coordinates' and isinstance(v, (list, dict)):
                            row_data[k] = str(v)
                        else:
                            row_data[k] = v
                    row_data['_geometry_type'] = geometry_type
                    row_data['_coordinates'] = coords

                    csv_data = CSVData.objects.create(
                        file_upload=self.file_upload,
                        row_number=index + 1,
                        data=row_data,
                        geometry_type=geometry_type,
                        coordinates=json.dumps(coords)
                    )
                    csv_data_objects.append(csv_data)
                except (ValueError, TypeError, SyntaxError) as e:
                    print(f"Error processing CSV row {index + 1}: {e}")
                    continue

            # Update file metadata
            self.file_upload.geometry_type = geometry_type if csv_data_objects else None
            self.file_upload.feature_count = len(csv_data_objects)
            self.file_upload.bounds = self._calculate_csv_bounds(csv_data_objects)
            self.file_upload.status = 'completed'
            self.file_upload.save()

            return {
                'success': True,
                'data_count': len(csv_data_objects),
                'geometry_type': geometry_type if csv_data_objects else None,
                'bounds': self.file_upload.bounds,
                'columns': list(df.columns)
            }

        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            raise
    
    def _process_shapefile(self):
        """Process Shapefile (ZIP) and extract data"""
        try:
            # Extract ZIP file
            temp_dir = tempfile.mkdtemp()
            
            with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Find .shp file
            shp_files = [f for f in os.listdir(temp_dir) if f.endswith('.shp')]
            if not shp_files:
                raise ValueError("No .shp file found in ZIP")
            
            shp_path = os.path.join(temp_dir, shp_files[0])
            
            # Read shapefile using geopandas
            gdf = gpd.read_file(shp_path, engine='pyogrio')
            
            if gdf.empty:
                raise ValueError("Shapefile is empty")
            
            # Process each feature
            from .models import ShapefileData
            shapefile_data_objects = []
            
            for index, row in gdf.iterrows():
                try:
                    geometry = row.geometry
                    if geometry is None:
                        continue
                    
                    # Convert geometry to coordinates
                    coords = self._geometry_to_coordinates(geometry)
                    
                    # Store feature data
                    attributes = row.drop('geometry').to_dict()
                    
                    shapefile_data = ShapefileData.objects.create(
                        file_upload=self.file_upload,
                        feature_id=index + 1,
                        geometry_type=geometry.geom_type,
                        coordinates=json.dumps(coords),
                        attributes=attributes
                    )
                    shapefile_data_objects.append(shapefile_data)
                    
                except Exception as e:
                    logger.warning(f"Error processing shapefile feature {index + 1}: {e}")
                    continue
            
            # Update file metadata
            self.file_upload.geometry_type = gdf.geometry.geom_type.iloc[0] if not gdf.empty else 'Unknown'
            self.file_upload.coordinate_system = str(gdf.crs) if gdf.crs else 'Unknown'
            self.file_upload.feature_count = len(shapefile_data_objects)
            self.file_upload.bounds = self._calculate_shapefile_bounds(shapefile_data_objects)
            self.file_upload.status = 'completed'
            self.file_upload.save()
            
            # Cleanup
            import shutil
            shutil.rmtree(temp_dir)
            
            return {
                'success': True,
                'data_count': len(shapefile_data_objects),
                'geometry_type': self.file_upload.geometry_type,
                'coordinate_system': self.file_upload.coordinate_system,
                'bounds': self.file_upload.bounds
            }
            
        except Exception as e:
            logger.error(f"Error processing shapefile: {e}")
            raise
    
    def _detect_coordinate_columns(self, df):
        """Detect latitude and longitude columns in CSV, or fallback to 'Coordinates' column for polygons"""
        columns = df.columns.str.lower()
        lat_patterns = ['lat', 'latitude', 'y', 'y_coord', 'ycoord']
        lon_patterns = ['lon', 'long', 'longitude', 'lng', 'x', 'x_coord', 'xcoord']

        lat_col = None
        lon_col = None

        # Usual detection
        for pattern in lat_patterns:
            if pattern in columns:
                lat_col = df.columns[columns == pattern][0]
                break
        for pattern in lon_patterns:
            if pattern in columns:
                lon_col = df.columns[columns == pattern][0]
                break

        # Fallback: if 'coordinates' column exists, use it
        if (not lat_col or not lon_col) and 'coordinates' in columns:
            return None, None  # Signal to use coordinates column

        return lat_col, lon_col

    def _geometry_to_coordinates(self, geometry):
        """Convert shapely geometry to coordinates"""
        if geometry.geom_type == 'Point':
            return [geometry.x, geometry.y]
        elif geometry.geom_type == 'LineString':
            return list(geometry.coords)
        elif geometry.geom_type == 'Polygon':
            return list(geometry.exterior.coords)
        elif geometry.geom_type == 'MultiPolygon':
            return [list(poly.exterior.coords) for poly in geometry.geoms]
        else:
            return list(geometry.coords)
    
    def _get_geometry_type_from_data(self, data):
        """Get geometry type from parsed data"""
        types = set(item.get('geometry_type', '') for item in data)
        if len(types) == 1:
            return list(types)[0]
        else:
            return 'Mixed'
    
    def _calculate_bounds(self, data):
        """Calculate bounds from parsed data"""
        min_lat, max_lat = float('inf'), float('-inf')
        min_lon, max_lon = float('inf'), float('-inf')
        
        for item in data:
            try:
                coords = json.loads(item.get('coordinates', '[]'))
                if isinstance(coords, list):
                    if item.get('geometry_type') == 'Point':
                        lon, lat = coords[0], coords[1]
                        min_lat = min(min_lat, lat)
                        max_lat = max(max_lat, lat)
                        min_lon = min(min_lon, lon)
                        max_lon = max(max_lon, lon)
                    elif item.get('geometry_type') == 'Polygon':
                        for coord in coords:
                            lon, lat = coord[0], coord[1]
                            min_lat = min(min_lat, lat)
                            max_lat = max(max_lat, lat)
                            min_lon = min(min_lon, lon)
                            max_lon = max(max_lon, lon)
            except:
                continue
        
        if min_lat != float('inf'):
            return {
                'min_lat': min_lat,
                'max_lat': max_lat,
                'min_lon': min_lon,
                'max_lon': max_lon
            }
        return {}
    
    def _calculate_csv_bounds(self, csv_data_objects):
        """Calculate bounds from CSV data"""
        min_lat, max_lat = float('inf'), float('-inf')
        min_lon, max_lon = float('inf'), float('-inf')
        
        for obj in csv_data_objects:
            try:
                coords = json.loads(obj.coordinates)
                lon, lat = coords[0], coords[1]
                min_lat = min(min_lat, lat)
                max_lat = max(max_lat, lat)
                min_lon = min(min_lon, lon)
                max_lon = max(max_lon, lon)
            except:
                continue
        
        if min_lat != float('inf'):
            return {
                'min_lat': min_lat,
                'max_lat': max_lat,
                'min_lon': min_lon,
                'max_lon': max_lon
            }
        return {}
    
    def _calculate_shapefile_bounds(self, shapefile_data_objects):
        """Calculate bounds from shapefile data"""
        min_lat, max_lat = float('inf'), float('-inf')
        min_lon, max_lon = float('inf'), float('-inf')
        
        for obj in shapefile_data_objects:
            try:
                coords = json.loads(obj.coordinates)
                if isinstance(coords[0], list):  # Polygon or LineString
                    for coord in coords:
                        lon, lat = coord[0], coord[1]
                        min_lat = min(min_lat, lat)
                        max_lat = max(max_lat, lat)
                        min_lon = min(min_lon, lon)
                        max_lon = max(max_lon, lon)
                else:  # Point
                    lon, lat = coords[0], coords[1]
                    min_lat = min(min_lat, lat)
                    max_lat = max(max_lat, lat)
                    min_lon = min(min_lon, lon)
                    max_lon = max(max_lon, lon)
            except:
                continue
        
        if min_lat != float('inf'):
            return {
                'min_lat': min_lat,
                'max_lat': max_lat,
                'min_lon': min_lon,
                'max_lon': max_lon
            }
        return {}

class FileConverter:
    """File conversion utilities"""
    
    @staticmethod
    def kml_to_csv(kml_data_objects, filename):
        """Convert KML data to CSV"""
        try:
            csv_data = []
            for kml_data in kml_data_objects:
                row_data = kml_data.get_all_fields_for_csv()
                csv_data.append(row_data)
            
            if not csv_data:
                raise ValueError("No data available for conversion")
            
            df = pd.DataFrame(csv_data)
            
            # Create response
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            df.to_csv(response, index=False, encoding='utf-8-sig')
            
            return response
            
        except Exception as e:
            logger.error(f"Error converting KML to CSV: {e}")
            raise
    
    @staticmethod
    def kml_to_shapefile(kml_data_objects, filename):
        """Convert KML data to Shapefile"""
        try:
            # Create temporary directory
            temp_dir = tempfile.mkdtemp()
            shapefile_path = os.path.join(temp_dir, f"{filename}.shp")
            
            # Prepare data for GeoDataFrame
            features = []
            for kml_data in kml_data_objects:
                try:
                    coords = json.loads(kml_data.coordinates)
                    
                    if kml_data.geometry_type == 'Point':
                        geom = Point(coords[0], coords[1])
                    elif kml_data.geometry_type == 'Polygon':
                        geom = Polygon(coords)
                    elif kml_data.geometry_type == 'LineString':
                        geom = LineString(coords)
                    else:
                        continue
                    
                    # Create feature with all available fields
                    feature = {
                        'geometry': geom,
                        'placemark_name': kml_data.placemark_name or '',
                        'kitta_number': kml_data.kitta_number or '',
                        'owner_name': kml_data.owner_name or '',
                        'geometry_type': kml_data.geometry_type,
                        'area_hectares': float(kml_data.area_hectares) if kml_data.area_hectares else None,
                        'area_sqm': float(kml_data.area_sqm) if kml_data.area_sqm else None,
                        'description': kml_data.description or '',
                        'snippet': kml_data.snippet or '',
                        'visibility': kml_data.visibility or '',
                        'open_status': kml_data.open_status or '',
                        'time_begin': kml_data.time_begin or '',
                        'time_end': kml_data.time_end or '',
                        'time_when': kml_data.time_when or '',
                        'altitude': kml_data.altitude or '',
                        'altitude_mode': kml_data.altitude_mode or '',
                        'tessellate': kml_data.tessellate or '',
                        'extrude': kml_data.extrude or '',
                        'style_id': kml_data.style_id or '',
                        'style_url': kml_data.style_url or '',
                        'address': kml_data.address or '',
                        'country_code': kml_data.country_code or '',
                        'administrative_area': kml_data.administrative_area or '',
                        'sub_administrative_area': kml_data.sub_administrative_area or '',
                        'locality': kml_data.locality or '',
                        'sub_locality': kml_data.sub_locality or '',
                        'thoroughfare': kml_data.thoroughfare or '',
                        'postal_code': kml_data.postal_code or '',
                        'phone_number': kml_data.phone_number or '',
                        'snippet': kml_data.snippet or '',
                        'atom_author': kml_data.atom_author or '',
                        'atom_link': kml_data.atom_link or '',
                        'xal_address_details': kml_data.xal_address_details or '',
                        'extended_data': json.dumps(kml_data.extended_data) if kml_data.extended_data else '',
                    }
                    features.append(feature)
                    
                except Exception as e:
                    logger.warning(f"Error processing feature for shapefile: {e}")
                    continue
            
            if not features:
                raise ValueError("No valid features found for shapefile export")
            
            # Create GeoDataFrame
            gdf = gpd.GeoDataFrame(features, crs='EPSG:4326')
            
            # Save to shapefile
            gdf.to_file(shapefile_path, engine='pyogrio')
            
            # Create zip file
            zip_path = os.path.join(temp_dir, f"{filename}.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in os.listdir(temp_dir):
                    if file.endswith(('.shp', '.shx', '.dbf', '.prj', '.cpg')):
                        file_path = os.path.join(temp_dir, file)
                        zipf.write(file_path, file)
            
            # Read zip file and create response
            with open(zip_path, 'rb') as f:
                zip_content = f.read()
            
            response = HttpResponse(zip_content, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename}.zip"'
            
            # Cleanup
            import shutil
            shutil.rmtree(temp_dir)
            
            return response
            
        except Exception as e:
            logger.error(f"Error converting KML to shapefile: {e}")
            # Cleanup on error
            if 'temp_dir' in locals():
                import shutil
                shutil.rmtree(temp_dir)
            raise
    
    @staticmethod
    def csv_to_kml(csv_data_objects, filename):
        """Convert CSV data to KML"""
        try:
            from xml.dom.minidom import Document
            
            doc = Document()
            kml_elem = doc.createElement('kml')
            kml_elem.setAttribute('xmlns', 'http://www.opengis.net/kml/2.2')
            doc.appendChild(kml_elem)
            
            document_elem = doc.createElement('Document')
            kml_elem.appendChild(document_elem)
            
            for csv_data in csv_data_objects:
                try:
                    data = csv_data.data
                    coords = json.loads(csv_data.coordinates)
                    
                    placemark = doc.createElement('Placemark')
                    
                    # Add name
                    name_elem = doc.createElement('name')
                    name_elem.appendChild(doc.createTextNode(str(data.get('name', f'Point {csv_data.row_number}'))))
                    placemark.appendChild(name_elem)
                    
                    # Add description
                    desc_elem = doc.createElement('description')
                    desc_text = ', '.join([f"{k}: {v}" for k, v in data.items() if k not in ['name', 'lat', 'lon', 'latitude', 'longitude']])
                    desc_elem.appendChild(doc.createTextNode(desc_text))
                    placemark.appendChild(desc_elem)
                    
                    # Add point geometry
                    point_elem = doc.createElement('Point')
                    coords_elem = doc.createElement('coordinates')
                    coords_elem.appendChild(doc.createTextNode(f"{coords[0]},{coords[1]}"))
                    point_elem.appendChild(coords_elem)
                    placemark.appendChild(point_elem)
                    
                    document_elem.appendChild(placemark)
                    
                except Exception as e:
                    logger.warning(f"Error processing CSV row for KML: {e}")
                    continue
            
            kml_str = doc.toprettyxml(indent='  ', encoding='utf-8')
            response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
            response['Content-Disposition'] = f'attachment; filename="{filename}.kml"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error converting CSV to KML: {e}")
            raise
    
    @staticmethod
    def csv_to_shapefile(csv_data_objects, filename):
        """Convert CSV data to Shapefile (supports Point and Polygon)"""
        try:
            # Create temporary directory
            temp_dir = tempfile.mkdtemp()
            shapefile_path = os.path.join(temp_dir, f"{filename}.shp")
            
            # Prepare data for GeoDataFrame
            features = []
            for csv_data in csv_data_objects:
                try:
                    coords = json.loads(csv_data.coordinates)
                    # Detect geometry type
                    if isinstance(coords[0], (float, int)) and isinstance(coords[1], (float, int)):
                        # Point geometry
                        geom = Point(coords[0], coords[1])
                    elif isinstance(coords[0], (list, tuple)) and len(coords[0]) == 2:
                        # Polygon geometry (list of [lon, lat] pairs)
                        geom = Polygon(coords)
                    else:
                        raise ValueError("Invalid coordinates format for shapefile export")
                    # Create feature with all data
                    feature = {
                        'geometry': geom,
                        **csv_data.data
                    }
                    features.append(feature)
                except Exception as e:
                    logger.warning(f"Error processing CSV row for shapefile: {e}")
                    continue
            if not features:
                raise ValueError("No valid features found for shapefile export")
            # Create GeoDataFrame
            gdf = gpd.GeoDataFrame(features, crs='EPSG:4326')
            # Save to shapefile
            gdf.to_file(shapefile_path, engine='pyogrio')
            # Create zip file
            zip_path = os.path.join(temp_dir, f"{filename}.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in os.listdir(temp_dir):
                    if file.endswith(('.shp', '.shx', '.dbf', '.prj', '.cpg')):
                        file_path = os.path.join(temp_dir, file)
                        zipf.write(file_path, file)
            # Read zip file and create response
            with open(zip_path, 'rb') as f:
                zip_content = f.read()
            response = HttpResponse(zip_content, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename}.zip"'
            # Cleanup
            import shutil
            shutil.rmtree(temp_dir)
            return response
        except Exception as e:
            logger.error(f"Error converting CSV to shapefile: {e}")
            # Cleanup on error
            if 'temp_dir' in locals():
                import shutil
                shutil.rmtree(temp_dir)
            raise
    
    @staticmethod
    def shapefile_to_kml(shapefile_data_objects, filename):
        """Convert Shapefile data to KML"""
        try:
            from xml.dom.minidom import Document
            
            doc = Document()
            kml_elem = doc.createElement('kml')
            kml_elem.setAttribute('xmlns', 'http://www.opengis.net/kml/2.2')
            doc.appendChild(kml_elem)
            
            document_elem = doc.createElement('Document')
            kml_elem.appendChild(document_elem)
            
            for shapefile_data in shapefile_data_objects:
                try:
                    attributes = shapefile_data.attributes
                    coords = json.loads(shapefile_data.coordinates)
                    
                    placemark = doc.createElement('Placemark')
                    
                    # Add name
                    name_elem = doc.createElement('name')
                    name_elem.appendChild(doc.createTextNode(str(attributes.get('name', f'Feature {shapefile_data.feature_id}'))))
                    placemark.appendChild(name_elem)
                    
                    # Add description
                    desc_elem = doc.createElement('description')
                    desc_text = ', '.join([f"{k}: {v}" for k, v in attributes.items() if k != 'name'])
                    desc_elem.appendChild(doc.createTextNode(desc_text))
                    placemark.appendChild(desc_elem)
                    
                    # Add geometry based on type
                    if shapefile_data.geometry_type == 'Point':
                        point_elem = doc.createElement('Point')
                        coords_elem = doc.createElement('coordinates')
                        coords_elem.appendChild(doc.createTextNode(f"{coords[0]},{coords[1]}"))
                        point_elem.appendChild(coords_elem)
                        placemark.appendChild(point_elem)
                    elif shapefile_data.geometry_type == 'LineString':
                        line_elem = doc.createElement('LineString')
                        coords_elem = doc.createElement('coordinates')
                        coords_str = ' '.join([f"{c[0]},{c[1]}" for c in coords])
                        coords_elem.appendChild(doc.createTextNode(coords_str))
                        line_elem.appendChild(coords_elem)
                        placemark.appendChild(line_elem)
                    elif shapefile_data.geometry_type == 'Polygon':
                        poly_elem = doc.createElement('Polygon')
                        outer_elem = doc.createElement('outerBoundaryIs')
                        linear_elem = doc.createElement('LinearRing')
                        coords_elem = doc.createElement('coordinates')
                        coords_str = ' '.join([f"{c[0]},{c[1]}" for c in coords])
                        coords_elem.appendChild(doc.createTextNode(coords_str))
                        linear_elem.appendChild(coords_elem)
                        outer_elem.appendChild(linear_elem)
                        poly_elem.appendChild(outer_elem)
                        placemark.appendChild(poly_elem)
                    
                    document_elem.appendChild(placemark)
                    
                except Exception as e:
                    logger.warning(f"Error processing shapefile feature for KML: {e}")
                    continue
            
            kml_str = doc.toprettyxml(indent='  ', encoding='utf-8')
            response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
            response['Content-Disposition'] = f'attachment; filename="{filename}.kml"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error converting shapefile to KML: {e}")
            raise
    
    @staticmethod
    def shapefile_to_csv(shapefile_data_objects, filename):
        """Convert Shapefile data to CSV"""
        try:
            csv_data = []
            for shapefile_data in shapefile_data_objects:
                try:
                    attributes = shapefile_data.attributes
                    coords = json.loads(shapefile_data.coordinates)
                    
                    # Add geometry information to attributes
                    row_data = {
                        'feature_id': shapefile_data.feature_id,
                        'geometry_type': shapefile_data.geometry_type,
                        'coordinates': json.dumps(coords),
                        **attributes
                    }
                    csv_data.append(row_data)
                    
                except Exception as e:
                    logger.warning(f"Error processing shapefile feature for CSV: {e}")
                    continue
            
            if not csv_data:
                raise ValueError("No data available for conversion")
            
            df = pd.DataFrame(csv_data)
            
            # Create response
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            df.to_csv(response, index=False, encoding='utf-8-sig')
            
            return response
            
        except Exception as e:
            logger.error(f"Error converting shapefile to CSV: {e}")
            raise

class FileExporter:
    """File export utility for different formats"""
    
    @classmethod
    def export_file(cls, file_upload, export_format):
        """Export file in different formats"""
        try:
            if export_format == 'csv':
                return cls._export_to_csv(file_upload)
            elif export_format == 'excel':
                return cls._export_to_excel(file_upload)
            elif export_format == 'json':
                return cls._export_to_json(file_upload)
            elif export_format == 'shapefile':
                return cls._export_to_shapefile(file_upload)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")
                
        except Exception as e:
            logger.error(f"Error exporting file {file_upload.id}: {e}")
            raise
    
    @classmethod
    def _export_to_csv(cls, file_upload):
        """Export file to CSV format"""
        if file_upload.file_type == 'kml':
            # Export KML data to CSV
            from .models import KMLFile, KMLData
            
            kml_file = KMLFile.objects.filter(
                user=file_upload.user,
                original_filename=file_upload.original_filename
            ).first()
            
            if kml_file:
                data = []
                for kml_data in kml_file.parsed_data.all():
                    data.append({
                        'Kitta Number': kml_data.kitta_number or '',
                        'Owner Name': kml_data.owner_name or '',
                        'Geometry Type': kml_data.geometry_type,
                        'Area (Hectares)': float(kml_data.area_hectares) if kml_data.area_hectares else '',
                        'Area (sqm)': float(kml_data.area_sqm) if kml_data.area_sqm else '',
                        'Description': kml_data.description or '',
                        'Placemark Name': kml_data.placemark_name or ''
                    })
                
                # Create CSV response
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}.csv"'
                
                if data:
                    writer = csv.DictWriter(response, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
                
                return response
        
        # For other file types, return original file
        response = HttpResponse(file_upload.file, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}"'
        return response
    
    @classmethod
    def _export_to_excel(cls, file_upload):
        """Export file to Excel format"""
        if file_upload.file_type == 'csv':
            # Convert CSV to Excel
            df = pd.read_csv(file_upload.file.path)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)
            
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}.xlsx"'
            return response
        
        # For other file types, return original file
        response = HttpResponse(file_upload.file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}"'
        return response
    
    @classmethod
    def _export_to_json(cls, file_upload):
        """Export file to JSON format"""
        if file_upload.file_type == 'csv':
            # Convert CSV to JSON
            df = pd.read_csv(file_upload.file.path)
            json_data = df.to_json(orient='records', indent=2)
            
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}.json"'
            return response
        
        # For other file types, return original file
        response = HttpResponse(file_upload.file, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}"'
        return response
    
    @classmethod
    def _export_to_shapefile(cls, file_upload):
        """Export file to Shapefile format"""
        if file_upload.file_type == 'kml':
            # Convert KML to Shapefile
            from .models import KMLFile, KMLData
            from .kml_utils import KMLExporter
            
            kml_file = KMLFile.objects.filter(
                user=file_upload.user,
                original_filename=file_upload.original_filename
            ).first()
            
            if kml_file:
                kml_data_list = []
                for kml_data in kml_file.parsed_data.all():
                    kml_data_list.append({
                        'kitta_number': kml_data.kitta_number,
                        'owner_name': kml_data.owner_name,
                        'geometry_type': kml_data.geometry_type,
                        'coordinates': json.loads(kml_data.coordinates),
                        'area_hectares': float(kml_data.area_hectares) if kml_data.area_hectares else None,
                        'area_sqm': float(kml_data.area_sqm) if kml_data.area_sqm else None,
                        'description': kml_data.description,
                        'placemark_name': kml_data.placemark_name
                    })
                
                # Create shapefile
                filename = os.path.splitext(file_upload.original_filename)[0]
                return KMLExporter.export_to_shapefile(kml_data_list, filename)
        
        # For other file types, return original file
        response = HttpResponse(file_upload.file, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{file_upload.original_filename}"'
        return response

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def log_download(user, file_upload, download_type, file_path, file_size, request):
    """Log file download"""
    from .models import DownloadLog
    
    DownloadLog.objects.create(
        user=user,
        kml_file=file_upload,  # This will need to be updated for general file uploads
        download_type=download_type,
        file_path=file_path,
        file_size=file_size,
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

def cleanup_temp_files(file_paths):
    """Clean up temporary files"""
    for file_path in file_paths:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            logger.warning(f"Error cleaning up temp file {file_path}: {e}") 